import os
import glob
import h5py
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime

PASTA_DADOS = "data"

ARQUIVO_RELATORIO_TXT = "relatorio_resumo_datasets.txt"
ARQUIVO_RELATORIO_XLSX = "relatorio_grupos_datasets.xlsx"

EXTENSOES_SUPORTADAS = ['.hdf', '.h5', '.csv', '.xlsx', '.xls', '.parquet']

REGRAS_CLASSIFICACAO = {
    'Powertrain & Motor (RPM, Torque, Potencia)': [
        'rpm', 'enginespeed', 'engine_speed', 'torque', 'trq', 'engine_load', 
        'coolant', 'engine_temp', 'oil_pressure', 'throttle', 'throttle_pos'
    ],
    'Transmissao & Troca de Marchas': [
        'gear', 'selected_gear', 'actual_gear', 'clutch', 'clutch_pos', 'transmission'
    ],
    'Frenagem & Pedais': [
        'brake', 'brake_pressure', 'brake_pos', 'pedal', 'accpedal', 'deceleration'
    ],
    'Dinamica Longitudinal (Velocidade & Aceleracao X)': [
        'vehiclespeed', 'speed', 'v_x', 'accel_x', 'long_accel', 'wheel_speed'
    ],
    'Dinamica Lateral & Curvas (IMU / Giroscopio)': [
        'accel_y', 'accel_z', 'lat_accel', 'gyro', 'gyroscope', 'pitch', 'roll', 
        'yaw', 'yaw_rate', 'steering', 'steering_angle', 'slip_angle', 'g_force', 'imu'
    ],
    'Consumo de Combustivel & Gases': [
        'fuel', 'fuelconsumption', 'fuel_rate', 'l/100km', 'km/l', 'diesel', 
        'gasoline', 'co2', 'emissions'
    ],
    'Bateria & Sistema Eletrico (EV / Hibrido)': [
        'soc', 'battery', 'battery_temp', 'current', 'voltage', 'power', 
        'energy', 'kwh', 'charging_status'
    ],
    'Navegacao, Rota & Elevacao (GPS)': [
        'latitude', 'longitude', 'lat', 'lon', 'altitude', 'elevation', 'gps', 
        'satellites', 'heading', 'direction', 'route', 'distance', 'track', 'tracks'
    ],
    'Gestao de Frota & Viagens (Trip Logs)': [
        'fleet', 'fleet_id', 'trip', 'trip_id', 'status', 'start_time', 'end_time'
    ],
    'Perfil de Veiculo & Especificacoes Tecnicas': [
        'driver', 'driver_id', 'vehicle_id', 'vehid', 'vin', 'model', 'brand', 
        'year', 'weight', 'generalized_weight[lb]', 'drive_wheels', 'engine_configuration'
    ]
}

def limpar_relatorios_antigos():
    padroes_para_remover = [
        ARQUIVO_RELATORIO_TXT,
        ARQUIVO_RELATORIO_XLSX,
        "relatorio_grupos_datasets_*.xlsx"
    ]
    for padrao in padroes_para_remover:
        arquivos_encontrados = glob.glob(padrao)
        for caminho in arquivos_encontrados:
            try:
                os.remove(caminho)
                print(f"[LIMPEZA] Arquivo antigo removido: '{caminho}'")
            except Exception as e:
                print(f"[AVISO] Nao foi possivel remover '{caminho}': {e}")

def extrair_colunas_hdf5(caminho):
    colunas = []
    try:
        with h5py.File(caminho, 'r') as hf:
            def callback(nome, obj):
                if isinstance(obj, h5py.Dataset):
                    nome_sinal = nome.split('/')[-1]
                    colunas.append(nome_sinal)
            hf.visititems(callback)
    except Exception:
        pass
    return colunas

def extrair_colunas_excel(caminho):
    colunas = []
    try:
        df_preview = pd.read_excel(caminho, nrows=2)
        colunas = list(df_preview.columns)
    except Exception:
        pass
    return colunas

def extrair_colunas_generico(caminho):
    nome_arq = os.path.basename(caminho)
    if nome_arq.startswith("~$") or nome_arq.startswith("relatorio_"):
        return []

    ext = os.path.splitext(caminho)[1].lower()
    try:
        if ext in ['.hdf', '.h5']:
            return extrair_colunas_hdf5(caminho)
        elif ext == '.csv':
            df_preview = pd.read_csv(caminho, nrows=2, sep=None, engine='python')
            return list(df_preview.columns)
        elif ext in ['.xlsx', '.xls']:
            return extrair_colunas_excel(caminho)
        elif ext == '.parquet':
            df_preview = pd.read_parquet(caminho)
            return list(df_preview.columns)
    except Exception:
        return []
    return []

def classificar_lista_colunas(lista_colunas):
    colunas_lower = [str(col).lower() for col in lista_colunas]
    pontuacao = {tema: 0 for tema in REGRAS_CLASSIFICACAO}

    for col in colunas_lower:
        for tema, palavras_chave in REGRAS_CLASSIFICACAO.items():
            if any(kw in col for kw in palavras_chave):
                pontuacao[tema] += 1

    temas_encontrados = [tema for tema, qtd in pontuacao.items() if qtd > 0]
    
    if not temas_encontrados:
        categoria_principal = "Nao identificado / Outros"
    elif len(temas_encontrados) == 1:
        categoria_principal = temas_encontrados[0]
    else:
        temas_ordenados = sorted(temas_encontrados, key=lambda t: pontuacao[t], reverse=True)
        categoria_principal = f"Misto (Dominante: {temas_ordenados[0]})"

    return categoria_principal, temas_encontrados

def construir_resumo_detalhado(nome_fonte, tipo_fonte, qtd_arquivos, tamanho_mb, colunas, temas):
    """Gera uma analise explicativa e contextual detalhada com base no conjunto de colunas e temas."""
    amostra_cols = ", ".join(colunas[:8]) if colunas else "Nenhuma coluna identificada"
    total_vars = len(colunas)
    
    detalhes_tematicos = []
    
    for t in temas:
        if 'Powertrain' in t:
            detalhes_tematicos.append("monitoramento de performance de motor e combustao (RPM, torque, carga)")
        elif 'Transmissao' in t:
            detalhes_tematicos.append("escalonamento e selecao de marchas")
        elif 'Frenagem' in t:
            detalhes_tematicos.append("dinamica de desaceleracao e acionamento de pedais")
        elif 'Longitudinal' in t:
            detalhes_tematicos.append("series temporais de velocidade do veiculo e aceleracao longitudinal")
        elif 'Lateral' in t:
            detalhes_tematicos.append("medicoes da IMU (aceleracoes laterais, giroscopio e angulo de estercamento/slip)")
        elif 'Consumo' in t:
            detalhes_tematicos.append("taxa de consumo de combustivel e emissao de efluentes")
        elif 'Bateria' in t:
            detalhes_tematicos.append("parametros eletricos de alta voltagem, estado de carga (SOC) e corrente")
        elif 'Navegacao' in t:
            detalhes_tematicos.append("posicionamento geografico via GPS, altitudes, rumo (heading) e trajetorias")
        elif 'Gestao' in t:
            detalhes_tematicos.append("historico de viagens (trips), status operacionais e IDs de acompanhamento")
        elif 'Perfil' in t:
            detalhes_tematicos.append("especificacoes estaticas de engenharia (peso, classe do veiculo, tipo de motor e tracao)")

    if detalhes_tematicos:
        foco_explicativo = "; ".join(detalhes_tematicos)
    else:
        foco_explicativo = "atributos gerais de dados nao diretamente mapeados no dicionario padrao"

    if "Pasta" in tipo_fonte:
        resumo = (
            f"Conjunto estruturado em pasta reunindo {qtd_arquivos} arquivo(s) (totalizando {tamanho_mb:.2f} MB) "
            f"e {total_vars} variavel(is) unicas. O perfil dos dados abrange: {foco_explicativo}. "
            f"Exemplo de atributos mapeados: [{amostra_cols}]."
        )
    else:
        resumo = (
            f"Arquivo de dados individual ({tamanho_mb:.2f} MB) estruturado com {total_vars} colunas/variaveis. "
            f"Contem registros focados em: {foco_explicativo}. "
            f"Amostra de colunas presentes: [{amostra_cols}]."
        )

    return resumo

def salvar_excel_formatado(dados_fontes, caminho_saida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo dos Datasets"

    headers = [
        "Dataset / Fonte", 
        "Tipo de Fonte", 
        "Qtd. Arquivos", 
        "Tamanho (MB)", 
        "Total Variaveis", 
        "Categoria Dominante", 
        "Subcategorias Detectadas", 
        "Resumo Detalhado e Perfil dos Dados"
    ]
    ws.append(headers)

    for item in dados_fontes:
        ws.append([
            item['Fonte_Dataset'],
            item['Tipo'],
            item['Qtd_Arquivos'],
            item['Tamanho_Total_MB'],
            item['Total_Variaveis'],
            item['Classificacao_Geral'],
            item['Temas_Detectados'],
            item['Resumo_Executivo']
        ])

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws.views.sheetView[0].showGridLines = True
    ws.row_dimensions[1].height = 28

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        ws.row_dimensions[row[0].row].height = 42
        for i, cell in enumerate(row):
            cell.font = data_font
            cell.border = thin_border
            if headers[i] in ["Qtd. Arquivos", "Total Variaveis"]:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif headers[i] in ["Tamanho (MB)"]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif headers[i] == "Resumo Detalhado e Perfil dos Dados":
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    col_widths = [28, 18, 14, 15, 16, 38, 45, 80]
    for i, col_letter in enumerate([openpyxl.utils.get_column_letter(c) for c in range(1, len(headers) + 1)]):
        ws.column_dimensions[col_letter].width = col_widths[i]

    try:
        wb.save(caminho_saida)
        return caminho_saida
    except PermissionError:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        novo_caminho = f"relatorio_grupos_datasets_{timestamp}.xlsx"
        wb.save(novo_caminho)
        print("\n[AVISO] A planilha original esta aberta no Excel.")
        print(f"[AVISO] O relatorio foi salvo em um novo arquivo: '{novo_caminho}'")
        return novo_caminho

def mapear_fontes_dados(pasta_base):
    if not os.path.exists(pasta_base):
        return []
    
    fontes = []
    for item in sorted(os.listdir(pasta_base)):
        if item.startswith('.') or item.startswith('~$') or item.startswith('relatorio_'):
            continue
            
        caminho_completo = os.path.join(pasta_base, item)
        
        if os.path.isdir(caminho_completo):
            arquivos = []
            for root, _, files in os.walk(caminho_completo):
                for file in files:
                    ext = os.path.splitext(file)[1].lower()
                    if ext in EXTENSOES_SUPORTADAS and not file.startswith('~$') and not file.startswith('relatorio_'):
                        arquivos.append(os.path.join(root, file))
            if arquivos:
                fontes.append({
                    'nome': item,
                    'tipo': 'Pasta / Grupo',
                    'arquivos': arquivos
                })
        else:
            ext = os.path.splitext(item)[1].lower()
            if ext in EXTENSOES_SUPORTADAS:
                fontes.append({
                    'nome': item,
                    'tipo': f'Arquivo ({ext.upper().replace(".", "")})',
                    'arquivos': [caminho_completo]
                })
                
    return fontes

def analisar_pasta_dados(pasta_base):
    limpar_relatorios_antigos()
    fontes = mapear_fontes_dados(pasta_base)
    
    if not fontes:
        print("=" * 75)
        print(f"ATENCAO: Nenhum dataset ou pasta encontrado dentro de '{pasta_base}'.")
        print("=" * 75)
        return

    dados_fontes = []
    conteudo_txt = []
    
    conteudo_txt.append("=" * 85)
    conteudo_txt.append("       RELATORIO DE ANÁLISE DETALHADA E CONTEXTUALIZADA DE DATASETS")
    conteudo_txt.append("=" * 85 + "\n")

    for fonte in fontes:
        nome_fonte = fonte['nome']
        tipo_fonte = fonte['tipo']
        arquivos = fonte['arquivos']

        colunas_fonte = set()
        tamanho_total_mb = 0

        for arq in arquivos:
            tamanho_mb = os.path.getsize(arq) / (1024 * 1024)
            tamanho_total_mb += tamanho_mb
            cols_arq = extrair_colunas_generico(arq)
            colunas_fonte.update(cols_arq)

        tamanho_total_mb = round(tamanho_total_mb, 2)
        col_fonte_lista = list(colunas_fonte)
        cat_fonte, temas_fonte = classificar_lista_colunas(col_fonte_lista) if col_fonte_lista else ("Sem colunas / Erro", [])
        temas_str = " | ".join(temas_fonte) if temas_fonte else "Nenhum"

        resumo_detalhado = construir_resumo_detalhado(
            nome_fonte=nome_fonte,
            tipo_fonte=tipo_fonte,
            qtd_arquivos=len(arquivos),
            tamanho_mb=tamanho_total_mb,
            colunas=col_fonte_lista,
            temas=temas_fonte
        )

        dados_fontes.append({
            'Fonte_Dataset': nome_fonte,
            'Tipo': tipo_fonte,
            'Qtd_Arquivos': len(arquivos),
            'Tamanho_Total_MB': tamanho_total_mb,
            'Total_Variaveis': len(col_fonte_lista),
            'Classificacao_Geral': cat_fonte,
            'Temas_Detectados': temas_str,
            'Resumo_Executivo': resumo_detalhado
        })

        conteudo_txt.append(
            f"DATASET / FONTE           : {nome_fonte}\n"
            f"-------------------------------------------------------------------------------------\n"
            f"- Tipo de Fonte           : {tipo_fonte}\n"
            f"- Qtd. de Arquivos        : {len(arquivos)}\n"
            f"- Tamanho Acumulado       : {tamanho_total_mb} MB\n"
            f"- Total de Variaveis      : {len(col_fonte_lista)}\n"
            f"- Categoria Dominante     : {cat_fonte}\n"
            f"- Subcategorias Mapeadas  : {temas_str}\n"
            f"- Resumo Detalhado        :\n"
            f"  {resumo_detalhado}\n"
            f"-------------------------------------------------------------------------------------\n\n"
        )

    if dados_fontes:
        with open(ARQUIVO_RELATORIO_TXT, 'w', encoding='utf-8') as f:
            f.write("\n".join(conteudo_txt))

        arquivo_salvo = salvar_excel_formatado(dados_fontes, ARQUIVO_RELATORIO_XLSX)

        print("=" * 75)
        print("ANALISE DETALHADA E CONTEXTUALIZADA CONCLUIDA COM SUCESSO!")
        print(f" Resumo em Texto : '{ARQUIVO_RELATORIO_TXT}'")
        print(f" Planilha Excel  : '{arquivo_salvo}'")
        print("=" * 75)

if __name__ == "__main__":
    analisar_pasta_dados(PASTA_DADOS)